import openpyxl

# Criar planilha(book)
book = openpyxl.Workbook()

# Como vizualizar páginas existentes
print(book.sheetnames)

# Como criar uma página
book.create_sheet('Frutas')

# Como selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Maçã', '2', 'R$15,60'])
frutas_page.append(['Pitaya', '10', 'R$32,97'])
frutas_page.append(['Maracujá', '2', 'R$60,42'])

# Salvar a planilha
book.save('C:\\Users\\diego.figueiro\\Desktop\\diegorepo\\Planilhas Automáticas\\Planilha de Compras.xlsx')
