import openpyxl

# Carregando o arquivo
book = openpyxl.load_workbook('C:\\Users\\diego.figueiro\\Desktop\\diegorepo\\Planilhas Automáticas\\Planilha de Compras.xlsx')

# Selecionando uma página
frutas_page = book['Frutas']

# Imprimindo os dados de cada linha

for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}')

# Alterando conteúodo específico de uma célula
for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        if cell.value == 'Banana':
            cell.value = 'Fruta 1'

# Salvar as alterações
book.save('C:\\Users\\diego.figueiro\\Desktop\\diegorepo\\Planilhas Automáticas\\Planilha de Compras.xlsx')